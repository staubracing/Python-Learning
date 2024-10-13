import os
import xmltodict
from openpyxl import Workbook

workbook = Workbook()
sheet = workbook.active
directory = """C://Users/Alexander.Chaddock/OneDrive - Olympus/Documents/CF Data/Avis"""

row = 1

for filename in os.listdir(directory):
    if filename.endswith(".xml"):
        with open(file=(directory+"/"+ filename),mode='r',encoding='utf-8') as file:
            XmlDictionary = xmltodict.parse(xml_input=file.read(),encoding='utf-8')
            
            #CreateHeader on first loop.
            if row == 1:
                column = 1
                for k in XmlDictionary['PartResults']['SummaryInfo'].keys():
                    sheet.cell(row = row, column = column, value = k)
                    column = column + 1

                for k in XmlDictionary['PartResults']['DetailInfo']['AttrMeasurements']['AttrMeasurement'].keys():
                    if k != 'Name':
                        sheet.cell(row = row, column=column, value = (str(XmlDictionary['PartResults']['DetailInfo']['AttrMeasurements']['AttrMeasurement']['Name']) +"-" + k))
                        column = column + 1

                for k in XmlDictionary['PartResults']['DetailInfo']['VarMeasurements']['VarMeasurement']:
                    for atr in k.keys():
                        if atr != 'Name':
                            sheet.cell(row = row, column=column,value=(k['Name']+"-"+atr))
                            column = column + 1
            
            row = row + 1

            #Fill in the appropriate data for each column
            if row > 1:
                column = 1
                for k in XmlDictionary['PartResults']['SummaryInfo'].keys():
                    sheet.cell(row = row, column = column, value = XmlDictionary['PartResults']['SummaryInfo'][k])
                    column = column + 1

                for k in XmlDictionary['PartResults']['DetailInfo']['AttrMeasurements']['AttrMeasurement'].keys():
                    if k != 'Name':
                        sheet.cell(row = row, column=column, value = XmlDictionary['PartResults']['DetailInfo']['AttrMeasurements']['AttrMeasurement'][k])
                        column = column + 1

                for k in XmlDictionary['PartResults']['DetailInfo']['VarMeasurements']['VarMeasurement']:
                    for atr in k.keys():
                        if atr != 'Name':
                            sheet.cell(row = row, column=column,value=k[atr])
                            column = column + 1

workbook.save(directory + "/ScriptOutput.xlsx")

print("complete!")